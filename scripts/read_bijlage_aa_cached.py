"""Lees cached values uit xlsm (uit laatste Excel-recalc indien aanwezig)."""

import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from pathlib import Path

XLSM = Path(__file__).parent.parent / "tests" / "references" / "bijlage-aa-sample-case1-slaapkamer-zuid.xlsm"

# data_only=True → toont gecachte values ipv formules
wb = openpyxl.load_workbook(XLSM, keep_vba=True, data_only=True)

print("=== Ruimte 1 — uitlees-cellen ===")
ws = wb["Ruimte 1"]
for coord, label in [
    ("B53", "θ_e max"),
    ("B55", "P_tr;ntr Voorgevel"),
    ("B56", "P_sol Voorgevel"),
    ("B57", "P_tr;gl Voorgevel"),
    ("B58", "Totaal per gevel Voorgevel"),
    ("B60", "P_int;calc"),
    ("B61", "P_v;calc"),
    ("B63", "Totaal koellastbijdrage"),
    ("B64", "Koelbehoefte verblijfsruimte"),
]:
    val = ws[coord].value
    print(f"  {coord} ({label:35s}) = {val!r}")

print("\n=== Projectgegevens en Resultaten — uitlees-cellen ===")
ws = wb["Projectgegevens en Resultaten"]
for coord, label in [
    ("B16", "f;iso (uit bouwjaarklasse)"),
    ("B33", "q_int;calc;zi"),
    ("B42", "Buitenluchttemp tmax,zi (Ruimte 1)"),
    ("B43", "Pint;calc (Ruimte 1)"),
    ("B44", "Pv;calc (Ruimte 1)"),
    ("B45", "Ptr;ntr (Ruimte 1)"),
    ("B46", "Psol (Ruimte 1)"),
    ("B47", "Ptr;gl (Ruimte 1)"),
    ("B49", "Koelbehoefte verblijfsruimte"),
    ("B51", "Minimaal benodigde koelcapaciteit"),
    ("B55", "Benodigde koelcapaciteit Ruimte 1"),
    ("B68", "Minimaal op te stellen koelvermogen opwekker"),
]:
    val = ws[coord].value
    print(f"  {coord} ({label:50s}) = {val!r}")
