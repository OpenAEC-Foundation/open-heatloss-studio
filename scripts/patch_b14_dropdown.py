"""Patch B14 in sample case 1 xlsm: integer 2020 → string 'vanaf 2015'.

De RVO-rekentool VBA `o_F_Iso` verwacht een dropdown-string als bouwjaar
('tot 1975' / '1975 t/m 1991' / '1992 t/m 2014' / 'vanaf 2015'). PM had
2020 als integer ingevuld → UDF matcht geen Case → return 0.
"""

import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from pathlib import Path

XLSM = Path(__file__).parent.parent / "tests" / "references" / "bijlage-aa-sample-case1-slaapkamer-zuid.xlsm"

wb = openpyxl.load_workbook(XLSM, keep_vba=True)
ws = wb["Projectgegevens en Resultaten"]

print(f"B14 before: {ws['B14'].value!r}")
ws["B14"] = "vanaf 2015"
print(f"B14 after:  {ws['B14'].value!r}")

# B15 (nageisoleerd) laat staan zoals het is — voor 'vanaf 2015' is er
# in de VBA UDF geen Ja/Nee branch, dus value maakt niet uit.
print(f"B15 (nageisoleerd):  {ws['B15'].value!r}")

wb.save(XLSM)
print(f"\nSaved: {XLSM}")
