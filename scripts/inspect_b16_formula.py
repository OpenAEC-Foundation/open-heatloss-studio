"""Lees de B16 ArrayFormula voor f_iso (bouwjaarklasse -> f_iso lookup)."""

import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from pathlib import Path

XLSM = Path(__file__).parent.parent / "tests" / "references" / "bijlage-aa-sample-case1-slaapkamer-zuid.xlsm"
wb = openpyxl.load_workbook(XLSM, keep_vba=True, data_only=False)

ws = wb["Projectgegevens en Resultaten"]

for coord in ["B16", "B18", "B31"]:
    cell = ws[coord]
    val = cell.value
    print(f"Cell {coord}:")
    print(f"  value (raw): {val!r}")
    if hasattr(val, "text"):
        print(f"  ArrayFormula text: {val.text!r}")
        print(f"  ArrayFormula ref:  {val.ref!r}")
    print()

# Tabellen sheet — zoek de f_iso lookup-tabel
print("=== Sheet 'Tabellen' — kolommen A/B/C, rij 1-50 ===")
ws_t = wb["Tabellen"]
print(f"Dimensions: {ws_t.dimensions}")
for row in ws_t.iter_rows(min_row=1, max_row=50, max_col=5, values_only=False):
    for cell in row:
        if cell.value is None:
            continue
        val = str(cell.value)
        if len(val) > 80:
            val = val[:77] + "..."
        print(f"  {cell.coordinate:6s} | {val}")
